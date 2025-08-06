import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment
from datetime import datetime, date
from app.models import Student
from django.conf import settings
import os


def create_may_collection_excel(filename="month_collection.xlsx", start=None, end=None):

    # Default date range: past 30 days
    if start is None or end is None:
        today = datetime.today().date()
        start = date(today.year, today.month - 1, today.day)
        end = today

    students = Student.objects.filter(joined_date__range=[start, end]).order_by("joined_date")

    # Prepare media directory
    dir_path = os.path.join(settings.BASE_DIR, "static", "media", "payment")
    os.makedirs(dir_path, exist_ok=True)
    filepath = os.path.join(dir_path, filename)

    # Excel setup
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "May Collection"

    max_receipts = max([s.receipts.count() for s in students], default=0)

    headers = ["DATE", "NAME", "COURSE", "DURATION", "TOTAL"]
    for i in range(1, max_receipts + 1):
        suffix = "ST" if i == 1 else "ND" if i == 2 else "RD" if i == 3 else "TH"
        headers.append(f"PAID_{i}{suffix}")
    headers += ["BALANCE", "CONTACT"]
    ws.append(headers)

    all_rows = []

    for col, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    for student in students:
        receipts = student.receipts.all().order_by('date')
        total_paid = sum(r.amount for r in receipts)
        balance = student.total_fee - total_paid if student.total_fee else ''
        balance_text = 'COMPLETED' if balance == 0 else balance

        receipt_cells = [f"{r.amount} ({r.date.strftime('%d/%m/%Y')})" for r in receipts]
        receipt_cells += [''] * (max_receipts - len(receipt_cells))

        row = [
            student.joined_date.strftime("%d/%m/%Y") if student.joined_date else '',
            student.name,
            student.course_name,
            student.duration,
            student.total_fee,
        ] + receipt_cells + [balance_text, student.phone]

        ws.append(row)
        all_rows.append({
            "Date": student.joined_date,
            "Name": student.name,
            "Course": student.course_name,
            "Total Fee": student.total_fee,
            "Total Paid": total_paid,
            "Balance": 0 if balance_text == "COMPLETED" else balance
        })
    if os.path.exists(filepath):
        os.remove(filepath)

    wb.save(filepath)

    # Create pivot table
    df = pd.DataFrame(all_rows)


    return filepath, df
