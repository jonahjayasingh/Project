import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from django.conf import settings  # Ensure this if used inside Django

def create_account_excel(transactions, name="account_excel"):
    # Prepare data
    records = []
    for t in transactions:
        records.append({
            "User": t.account.user.username,
            "Date": t.timestamp.strftime("%Y-%m-%d"),
            "Type": t.transaction_type.title(),
            "Amount": t.amount,
            "Description": t.description,
            "account_balance":t.account_balance
        })

    df = pd.DataFrame(records)
    if df.empty:
        return None, pd.DataFrame(columns=["No transactions found."])

    df.sort_values(by=["User", "Date"], inplace=True)
    df["Credit"] = df.apply(lambda row: row["Amount"] if row["Type"] == "Spend" else 0, axis=1)
    df["Debit"] = df.apply(lambda row: row["Amount"] if row["Type"] == "Received" else 0, axis=1)

    # Compute running balance per user
    df["Balance"] = 0
    for user in df["User"].unique():
        mask = df["User"] == user
        df.loc[mask, "Balance"] = df.loc[mask, "Credit"].cumsum() - df.loc[mask, "Debit"].cumsum()

    # Replace zeros with empty strings
    df["Credit"] = df["Credit"].replace(0, "")
    df["Debit"] = df["Debit"].replace(0, "")
    df["Balance"] = df["Balance"].replace(0, "")
    df["Balance"] = df["account_balance"].apply(lambda x: "" if x == "" else abs(x))

    df = df[["Date", "Description", "Credit", "Debit", "Balance"]]

    # Add grand total row
    total_row = pd.DataFrame([{
        "Date": "GRAND TOTAL",
        "Description": "",
        "Credit": df["Credit"].replace("", 0).sum(),
        "Debit": df["Debit"].replace("", 0).sum(),
        "Balance": df["Balance"].replace("", 0).iloc[-1]
    }])
    df_excel = pd.concat([df, total_row], ignore_index=True)

    # Save Excel file
    dir_path = os.path.join(settings.BASE_DIR, "static", "media", "account")
    os.makedirs(dir_path, exist_ok=True)
    file_path = os.path.join(dir_path, f"{name}.xlsx")
    df_excel.to_excel(file_path, index=False, engine='openpyxl')

    # Format the Excel sheet
    wb = load_workbook(file_path)
    ws = wb.active

    header_font = Font(bold=True)
    total_font = Font(bold=True, color="000000")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rupee_format = '₹ #,##0.00;₹ #,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col_letter].width = max_len + 6

        for cell in col:
            cell.alignment = center_align
            cell.border = thin_border
            if cell.row == 1:
                cell.font = header_font
            elif cell.row == ws.max_row:
                cell.font = total_font
            if cell.column in [3, 4, 5] and isinstance(cell.value, (int, float)):
                cell.number_format = rupee_format

    if os.path.exists(file_path):
        os.remove(file_path)
    wb.save(file_path)

    return os.path.basename(file_path), df_excel
