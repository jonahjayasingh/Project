from ast import Sub
from django.shortcuts import render
from django.contrib.auth.models import Permission, User
from django.contrib import messages
from django.db.models import Sum, F, ExpressionWrapper, FloatField
from django.shortcuts import redirect
from django.utils import timezone
from django.contrib.auth import authenticate,login, logout
from .models import *
from .customOperations.GC import generate_certificate
from .customOperations.BC import get_bond_certificate
from .customOperations.Receipt import generate_receipt
from django.conf import settings
import os
from django.http import FileResponse, Http404, HttpResponse
from datetime import datetime,date
import shutil
import pandas as pd
from .customOperations.Payment import create_may_collection_excel
from .customOperations.account_excel import create_account_excel
from django.core.mail import EmailMessage
from django.contrib.auth.decorators import login_required
from django.conf import settings
from staff.models import Work,Attendance
from openpyxl import Workbook

def home(request):
    return render(request,"home.html")

@login_required
def dashboard(request):

    try:
        students = Student.objects.all().order_by('joined_date').reverse()
        students = [student for student in students if student.certificate_id == None]
        faculty = User.objects.filter(is_staff=True , is_superuser=False)
        courses = CourseName.objects.all()
    except Exception as e:

        students = []
        faculty = []
        courses = []
    return render(request,"dashbord.html",{"students":students,"faculty":faculty,"courses":courses})

def user_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_superuser:
                login(request, user)
                return redirect("app:dashboard")
            elif user.is_staff:
                login(request, user)
                return redirect("staff:staff_dashboard")
            else:
                messages.error(request, "You are not authorized to login")
                return redirect("app:login")
        else:
            messages.error(request, "Invalid username or password")
    return render(request,"login.html")

def user_register(request):
    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        if password != confirm_password:
           messages.error(request,"Password does not match")
           return render(request,"register.html")
        if User.objects.filter(username=username).exists():
            messages.error(request,"Username already exists")
            return render(request,"register.html")
        if User.objects.filter(email=email).exists():
            messages.error(request,"Email already exists")
            return render(request,"register.html")
        user = User.objects.create_user(username=username,email=email,password=password)
        user.save()
        return redirect("app:login")
    return render(request,"register.html")

@login_required
def user_logout(request):
    user = User.objects.get(id=request.user.id)
    if user.is_superuser:
        try:
            if os.path.exists("certificategenrator/static/media/certificates"):
                shutil.rmtree("certificategenrator/static/media/certificates")
            if os.path.exists("certificategenrator/static/media/receipt"):
                shutil.rmtree("certificategenrator/static/media/receipt")
            if os.path.exists("certificategenrator/static/media/work"):
                shutil.rmtree("certificategenrator/static/media/work")
            if os.path.exists("certificategenrator/static/media/attendance"):
                shutil.rmtree("certificategenrator/static/media/attendance")
            if os.path.exists("certificategenrator/static/media/account"):
                shutil.rmtree("certificategenrator/static/media/account")
            if os.path.exists("certificategenrator/static/media/payment"):
                shutil.rmtree("certificategenrator/static/media/payment")
        except Exception as e:
            pass
    logout(request)

    return redirect("app:home")

def get_in_touch(request):
    if request.method == "POST":
        name = request.POST.get("name")
        phone = request.POST.get("phone")
        message = request.POST.get("message")
        get_in_touch = GetInTouch(name=name,phone=phone,message=message)
        get_in_touch.save()
        return redirect("app:home")
    return render(request,"get_in_touch.html")


@login_required
def add_student(request):
    if request.method == "POST":
        name = request.POST.get("name")
        phone = request.POST.get("phone")
        email = request.POST.get("email")
        course_name = request.POST.get("course_name")
        duration = request.POST.get("duration")
        total_fee = request.POST.get("total_fee")
        date = request.POST.get("date")
        dob = request.POST.get("dob")
        faculty = request.POST.get("faculty_name")
        pincode = request.POST.get("pincode")
        address = request.POST.get("address")
        qualification = request.POST.get("qualification")
        marital_status = request.POST.get("marital_status")
        online = request.POST.get("online")
        online = True if online == "on" else False
        const_value = Constant_Value.objects.get(id=1)
        const_value.registration_no +=1
        const_value.save()
        registration_no = const_value.registration_no
        faculty_name = User.objects.get(id =faculty)

        student = Student(name=name,phone=phone,email=email,registration_no=registration_no,course_name=course_name,duration=duration,total_fee=total_fee,joined_date=date,faculty_name=faculty_name,pincode=pincode,address=address,qualification=qualification,marital_status=marital_status,online=online,dob=dob)
        student.save()
        re = request.POST.get("redirect")

        if re:
            return redirect("staff:add_student")
        else:
            return redirect("app:dashboard")
    return redirect("app:dashboard")
@login_required
def edit_student(request):
    if request.method == "POST":
        name = request.POST.get("name")
        phone = request.POST.get("phone")
        email = request.POST.get("email")
        course_name = request.POST.get("course_name")
        duration = request.POST.get("duration")
        total_fee = request.POST.get("total_fee")
        faculty_name = request.POST.get("faculty_name")
        pincode = request.POST.get("pincode")
        address = request.POST.get("address")
        qualification = request.POST.get("qualification")
        marital_status = request.POST.get("marital_status")
        online = request.POST.get("online")
        online = True if online == "on" else False
        check = request.POST.get("redirect")
        dob = request.POST.get("dob")

        student = Student.objects.get(id=request.POST.get("student_id"))
        student.name = name
        student.phone = phone
        student.email = email
        student.course_name = course_name
        student.dob =dob
        student.duration = duration
        student.total_fee = total_fee
        student.faculty_name = faculty_name
        student.pincode = pincode
        student.address = address
        student.qualification = qualification
        student.marital_status = marital_status
        student.online = online
        student.save()

        messages.success(request, "Student details updated successfully")

        if check == "ss":
            return redirect("staff:single_student",id=student.id)
        return redirect("app:single_student",student_id=student.id)
    return redirect("app:dashboard")
@login_required
def show_certificate(request, location):
    try:
        filename = os.path.basename(location)
        pdf_path = os.path.join(settings.BASE_DIR,  "static",'media', 'certificates', filename)

        if not os.path.exists(pdf_path):
            raise Http404("Certificate not found")

        pdf_url = f"/media/certificates/{filename}"


    except Exception as e:

        raise Http404("Invalid certificate request")

    return render(request, "show_certificate.html", {
        "pdf_url": pdf_url,
        "filename": filename,
        'doc_type': 'certificate'
    })


@login_required
def search_alumni(request):
    if request.method == "POST":
        search_query = request.POST.get("search")
        search_by = request.POST.get("search_by")
        if search_by == "name":
            students = Student.objects.filter(certificate_id__isnull=False).filter(name__icontains=search_query)
        elif search_by == "phone":
            students = Student.objects.filter(certificate_id__isnull=False).filter(phone__icontains=search_query)
        elif search_by == "email":
            students = Student.objects.filter(certificate_id__isnull=False).filter(email__icontains=search_query)
        elif search_by == "registration_no":
            students = Student.objects.filter(certificate_id__isnull=False).filter(registration_no__icontains=search_query)
        return render(request,"oldStudents.html",{"alumni":students})
    else:
        students = Student.objects.filter(certificate_id__isnull=False)
        return render(request,"oldStudents.html",{"alumni":students})


@login_required
def get_certificate(request):
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        student = Student.objects.get(id=student_id)

        completed_date_str = request.POST.get("completed_date")
        try:
            completed_date = datetime.strptime(completed_date_str, "%Y-%m-%d").date()
        except:
            pass
        certificate_type = request.POST.get("certificate_type")
        range_date = request.POST.get("dateRange")
        project_title = request.POST.get("project_title")
        check = request.POST.get("redirect")

        if student.certificate_id is None:
            const_value = Constant_Value.objects.get(id=1)
            current_year = datetime.now().year


            latest_student = Student.objects.filter(certificate_id__isnull=False).order_by("-completed_date").first()


            has_current_year_cert = Student.objects.filter(
                certificate_id__isnull=False,
                completed_date__year=current_year
            ).exists()


            if not has_current_year_cert:
                const_value.certificate_id = 1

            else:
                const_value.certificate_id += 1

            const_value.save()

            certificate_id = const_value.certificate_id
            student.certificate_id = f'AL/IT-{current_year}/{certificate_id}'
            student.certificate_type = certificate_type
            student.completed_date = completed_date
            if range_date:
                student.date_range = range_date
            if project_title:
                student.project_title = project_title
            student.save()

        if certificate_type == "course" or student.certificate_type == "course":
            certificate_data = {
                'name': student.name,
                'course_name': student.course_name,
                'duration': student.duration,
                "certificate_id": f"Certificate No: {student.certificate_id}",
                "completed_date": student.completed_date
            }
            res = generate_certificate(certificate_data)
        else:
            certificate_data = {
                'name': student.name,
                'course_name': student.course_name,
                'duration': student.date_range,
                "certificate_id": student.certificate_id,
                "date_range": student.date_range,
                "completed_date": student.completed_date,
                "project_title": student.project_title
            }
            res = get_bond_certificate(certificate_data)

        if check == "ss":
            return redirect("staff:certificate", location=res)
        else:
            return redirect("app:show_certificate", location=res)

    return redirect("app:dashboard")



def send_email_with_pdf(pdf_path):
    id = pdf_path.split("_")[-1].split(".")[0]
    filename = pdf_path.split("\\")[-1]
    student = Student.objects.get(id=id)
    receipt = Receipt.objects.filter(student=student).order_by("-id")[0]

    subject = f'Recipt No: {receipt.reg_no} for {student.name} '
    body = f'Student Name: {student.name}\nCourse Name: {student.course_name}\nAmount: {receipt.amount}\nDate: {receipt.date.strftime("%d %B, %Y")}\n'
    from_email = settings.EMAIL_HOST_USER
    to_email = [settings.EMAIL_RECIVER ]

    email = EmailMessage(subject, body, from_email, to_email)

    with open(pdf_path, 'rb') as f:
        email.attach(f'{filename}', f.read(), 'application/pdf')

    email.send()

@login_required
def show_receipt(request, location):
    try:

        filename = os.path.basename(location)
        pdf_path = os.path.join(settings.BASE_DIR, "static",'media', 'receipt', filename)
        if not os.path.exists(pdf_path):
            raise Http404("Receipt not found")

        pdf_url = f"/media/receipt/{filename}"

    except Exception as e:
        raise Http404("Invalid receipt request")
    send_email_with_pdf(pdf_path)
    return render(request, "show_certificate.html", {
        "pdf_url": pdf_url,
        "filename": filename
    })

@login_required
def add_receipt(request):
    if request.method == "POST":
        const_value = Constant_Value.objects.get(id=1)
        const_value.receipt_reg_no += 1
        const_value.save()
        receipt_reg_no = const_value.receipt_reg_no
        student_id = request.POST.get("student_id")
        name = request.POST.get("name")
        course_name = request.POST.get("course_name")
        amount = request.POST.get("amount")
        date = request.POST.get("date")
        f_date = datetime.strptime(date, "%Y-%m-%d").date()
        f_date = f_date.strftime("%d %B, %Y")
        check = request.POST.get("redirect")

        student = Student.objects.get(id=request.POST.get("student_id"))



        receipt_data = {
            'name': name,
            "student_id":student.id,
            'course_name': course_name,
            'rupees': amount,
            "date":f_date,
            "reg_no":str(receipt_reg_no)

        }



        receipt = Receipt(user=request.user,name=name, student=student ,course_name=course_name,amount=amount,date=date,reg_no=str(receipt_reg_no))
        if  int(amount)  > student.total_fee:
            messages.error(request, f"Payment exceeds total fee ({student.total_fee}). You are trying to pay {amount}.")
            return redirect("app:single_student", student_id=student.id)
        else:

            if Payment.objects.filter(user=student).exists():
                payment = Payment.objects.get(user=student)
                if payment.paid_ammount + int(amount) <= student.total_fee:
                    payment.paid_ammount += int(amount)
                    payment.save()
                else:
                    messages.error(
                        request,
                        f"Payment exceeds total fee ({student.total_fee}). Already paid: {payment.paid_ammount}."
                    )
                    return redirect("app:single_student", student_id=student.id)
            else:
                payment = Payment(user=student,paid_ammount=amount)
            receipt.save()
            payment.save()
        res = generate_receipt(receipt_data)

        if check == "ss":
            return redirect("staff:receipt", location=res)
        return redirect("app:show_receipt",location=res)
    return redirect("app:dashboard")

@login_required
def search_student(request):
    if request.method == "POST":
        search_by = request.POST.get("search_by")
        search_query = request.POST.get("search")
        re= request.POST.get("redirect")
        if search_by == "name":
            students = Student.objects.filter(certificate_id__isnull=True).filter(name__icontains=search_query)
        elif search_by == "phone":
            students = Student.objects.filter(certificate_id__isnull=True).filter(phone__icontains=search_query)
        elif search_by == "email":
            students = Student.objects.filter(certificate_id__isnull=True).filter(email__icontains=search_query)
        elif search_by == "registration_no":
            students = Student.objects.filter(certificate_id__isnull=True).filter(registration_no__icontains=search_query)

        if re:
            try:
                permission = StaffPermission.objects.get(user=request.user)
            except StaffPermission.DoesNotExist:
                permission = None

            content ={
                "students":students,
                "permission":permission,
                "faculty":User.objects.filter(is_staff=True , is_superuser=False),
                "courses": CourseName.objects.all()
            }
            return render(request,"staff/add_Student.html",content)
        else:
            try:
                faculty = User.objects.filter(is_staff=True , is_superuser=False)
                courses = CourseName.objects.all()
            except Exception as e:
                faculty = []
                courses = []
            return render(request,"dashbord.html",{"students":students,"faculty":faculty,"courses":courses})
    return redirect("app:dashboard")

@login_required
def delete_user(request):
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        user = User.objects.get(id=user_id)
        user.delete()
        return redirect("app:staff")
    return redirect("app:staff")

def single_student(request,student_id):
    student = Student.objects.get(id=student_id)
    faculty = User.objects.filter(is_staff=True ,is_superuser=False)
    course = CourseName.objects.all()
    if Payment.objects.filter(user=student).exists():
        payment = Payment.objects.get(user=student)
    else:
        payment = None
    return render(request,"single_student.html",{"student":student,"faculty":faculty,"payment":payment,"courses":course})

@login_required
def delete_student(request,student_id):
    student = Student.objects.get(id=student_id)
    student.delete()
    return redirect("app:dashboard")

@login_required
def payment(request):
    year = datetime.now().year
    month = datetime.now().month
    students = Student.objects.filter(joined_date__year=year,joined_date__month=month).order_by('joined_date')
    payments = Payment.objects.filter(user__in=students)
    total_amount = payments.aggregate(total=Sum('paid_ammount'))['total'] or 0
    total_fee = students.aggregate(total=Sum('total_fee'))['total'] or 0
    if request.method == "POST":
        date = request.POST.get("dateRange")
        str_d, end_d = date.strip().split(" to ")
        start_date = datetime.strptime(str_d, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_d, "%Y-%m-%d").date()
        students = Student.objects.filter(joined_date__range=(start_date, end_date)).order_by('joined_date').annotate(total_paid=Sum('payment__paid_ammount')).annotate(
            pending=ExpressionWrapper(
                F('total_fee') - F('total_paid'),
                output_field=FloatField()
            )
        ).order_by('joined_date')


        payments = Payment.objects.filter(user__in=students)
        total_amount = payments.aggregate(total=Sum('paid_ammount'))['total'] or 0
        total_fee = students.aggregate(total=Sum('total_fee'))['total'] or 0
        content = {
            "students":students,
            "total_amount":total_amount,
            "total_fee":total_fee,
            "pending_amount":total_fee - total_amount,
            "show_table":True,
            "date":date
        }
        return render(request,"payment.html",content)

    content = {
        "students":students,
        "total_amount":total_amount,
        "total_fee":total_fee,
        "pending_amount":total_fee - total_amount,


    }
    return render(request,"payment.html",content)


@login_required
def staff(request):
    staff_users = User.objects.filter(is_staff=True, is_superuser=False)
    regular_users = User.objects.filter(is_staff=False, is_superuser=False)

    permissions = StaffPermission.objects.filter(user__in=staff_users).select_related('user')
    staff_details = StaffDetails.objects.filter(user__in=staff_users).select_related('user')

    details_dict = {detail.user_id: detail for detail in staff_details}
    staff_with_permissions = []
    staff_details = StaffDetails.objects.all()
    for user in staff_users:
        try:
            user_permissions = permissions.get(user=user)
            user_details = details_dict.get(user.id)

            staff_with_permissions.append({
                "user": user,
                "permissions": user_permissions,
                "staff": user_details
            })
        except StaffPermission.DoesNotExist:
            continue

    context = {
        "staff_with_permissions": staff_with_permissions,
        "user_reg": regular_users,
        "staff_details":staff_details
    }
    return render(request, "staff.html", context)
@login_required
def update_permission(request):
    if request.method == "POST":
        id = request.POST.get("user_id")
        rp= request.POST.get("receipt_permission")
        cp = request.POST.get("certificate_permission")
        sp = request.POST.get("student_permission")

        permission = StaffPermission.objects.get(user= id)
        permission.receipt_permission = rp
        permission.certificate_permission = cp
        permission.student_permission = sp
        permission.save()
        return redirect("app:staff")

    return redirect("app:staff")

@login_required
def make_staff(request):
    if request.method == "POST":
        id = request.POST.get("user_id")
        user = User.objects.get(id = id)
        if StaffPermission.objects.filter(user=user).exists():
            StaffPermission.objects.filter(user=user).delete()
        else:
            StaffPermission(user=user).save()
        user = User.objects.get(id = id)
        user.is_staff = not user.is_staff
        user.save()

        return redirect("app:staff")
    return redirect("app:staff")



@login_required
def account(request):
    # Handle balance creation
    if request.method == "POST" and 'amount' in request.POST:
        try:
            amount = int(request.POST.get("amount"))
            if amount <= 0:
                messages.error(request, "Amount must be positive")
                return redirect('account')

            Account(user=request.user,balance=amount).save()

        except ValueError:
            messages.error(request, "Invalid amount entered")
            return redirect('account')

    # Prepare account data
    try:
        account = Account.objects.get(user=request.user)
        transactions = Transaction.objects.filter(account=account).order_by("-timestamp")[:10]

        # Calculate monthly total (current month)
        monthly_total = Transaction.objects.filter(
            account=account,
            transaction_type="received",
            timestamp__year=timezone.now().year,
            timestamp__month=timezone.now().month
        ).aggregate(total=Sum('amount'))["total"] or 0

        # Calculate total spent (all withdrawals)
        total_spent = Transaction.objects.filter(
            account=account,
            transaction_type="spend",
        ).aggregate(total=Sum('amount'))['total'] or 0

    except Account.DoesNotExist:
        account = None
        transactions = None
        monthly_total = 0
        total_spent = 0
    day = str(timezone.now()).split(" ")
    day = day[0].split("-")[2]
    context = {
        "balance": account,
        "transactions": transactions,
        "monthly_total": monthly_total,
        "total_spent": total_spent,
        "days":int(day)-1
    }

    return render(request, "account.html", context)

@login_required
def transaction(request):
    if request.method == "POST":
        amount = request.POST.get("amount")
        description = request.POST.get("description")
        transaction_type = request.POST.get("transaction_type")
        user = User.objects.get(id=request.user.id)
        account = Account.objects.get(user=user)

        if transaction_type == "spend":
            account.balance -= int(amount)

            transactions = Transaction.objects.create(account=account,amount=amount,description=description,transaction_type="spend",account_balance = account.balance)
            transactions.save()
            account.save()

        else:

            account.balance += int(amount)
            transactions = Transaction.objects.create(account=account,amount=amount,description=description,transaction_type="received",account_balance = account.balance)
            transactions.save()
            account.save()
    return redirect("app:account")


import time
@login_required
def display_excel(request):
    if request.method == 'POST':
        date_range = request.POST.get('dateRange')

        if "to" in date_range:
            start_date, end_date = date_range.split(' to ')
            transactions = Transaction.objects.filter(
                timestamp__date__range=[start_date, end_date]
            )
        else:

             transactions = Transaction.objects.filter(timestamp__date=date_range)

        excel_file, pivot_table = create_account_excel(transactions, "account_report")

        if excel_file is None:
            return redirect("app:account")
        try:
            filename = os.path.basename(excel_file)
            pdf_path = os.path.join(settings.BASE_DIR,  "static",'media', 'account', filename)
            if not os.path.exists(pdf_path):
                raise Http404("Certificate not found")

            pdf_url = f"/media/account/{filename}"


        except Exception as e:

            raise Http404("Invalid certificate request")
        time.sleep(20)
        return render(request, 'excel.html', {
            'pivot_table': pivot_table.to_html(classes='table table-bordered', index=False),
            'excel_file': pdf_url,
            "title" :"Account Transaction Report"
        })
    return redirect("app:account")
@login_required
def display_payment_excel(request):
    if request.method == "POST":
        date = request.POST.get("date")
        if date != "":
            str_d, end_d = date.strip().split(" to ")
            start_date = datetime.strptime(str_d, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_d, "%Y-%m-%d").date()
            excel_file, pivot_table = create_may_collection_excel(start=start_date, end=end_date)

        else:
            excel_file, pivot_table = create_may_collection_excel()

        try:
            filename = os.path.basename(excel_file)
            pdf_path = os.path.join(settings.BASE_DIR,  "static",'media', 'payment', filename)
            if not os.path.exists(pdf_path):
                raise Http404("Certificate not found")

            pdf_url = f"/media/payment/{filename}"


        except Exception as e:

            raise Http404("Invalid certificate request")

        return render(request, 'excel.html', {
            'pivot_table': pivot_table.to_html(classes='table table-bordered', index=False),
            'excel_file': pdf_url,
            "title" :"Payment Report"
        })
    return redirect("app:payment")

@login_required
def download_student_details(request):
    students = Student.objects.all()
    df = pd.DataFrame(list(students.values()))
    file_path = "static/media/data/student_details.xlsx"
    directory = os.path.dirname(file_path)

    if not os.path.exists(directory):
        os.makedirs(directory)

    df.to_excel(file_path, index=False)
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='student_details.xlsx')


@login_required
def update_staff(request):
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        phone = request.POST.get("phone")
        designation = request.POST.get("designation")
        jd = request.POST.get("joining_date")
        DOB = request.POST.get("dob")
        image = request.FILES.get("image")
        rd = request.POST.get("resignation_date")
        e_id = request.POST.get("employee_id")
        salary = request.POST.get("salary")
        pincode = request.POST.get("pincode")
        address = request.POST.get("address")
        document = request.POST.get("documents")
        print(image)
        if rd == "":
            user = User.objects.get(id = user_id)
            if StaffDetails.objects.filter(email= user.email).exists():
                staff_details = StaffDetails.objects.get(name= user.username)
                try:
                    if image:
                        if staff_details.photo and staff_details.photo.name:
                            if os.path.isfile(staff_details.photo.path):
                                os.remove(staff_details.photo.path)
                        staff_details.photo = image
                except (OSError, ValueError):
                    pass



                staff_details.phone = phone
                staff_details.address = address
                staff_details.pincode = pincode
                staff_details.joining_date = jd
                staff_details.designation = designation
                staff_details.dob = DOB
                staff_details.salary = salary
                staff_details.employee_id = e_id
                staff_details.documents = document

                staff_details.save()
            else:
                StaffDetails(
                    user = User.objects.get(id = user_id),
                    phone = phone,
                    address = address,
                    pincode= pincode,
                    joining_date = jd,
                    designation = designation,
                    dob = DOB,
                    salary= salary,
                    employee_id = e_id,
                    documents= document,
                    photo=image

                ).save()

        else:
            user = User.objects.get(id = user_id)
            staff_details = StaffDetails.objects.get(name= user.username)
            staff_details.resignation_date = rd
            staff_details.save()


        return redirect("app:staff")
    return redirect("app:staff")

@login_required
def add_course(request):
    if request.method == "POST":
        course = request.POST.get("course_name")
        course = str(course).title()
        if course == "":
            return redirect("app:course")
        elif CourseName.objects.filter(course_name=course).exists():
            return redirect("app:course")

        CourseName(course_name=course).save()
    content={
        "courses": CourseName.objects.all()
    }
    return render(request,"course.html",content)

@login_required
def delete_course(request,id):
    course = CourseName.objects.get(id=id)
    course.delete()
    return redirect("app:course")

@login_required
def edit_course(request):
    if request.method == "POST":
        id = request.POST.get("id")
        course_name = str(request.POST.get("course_name")).title()
        course = CourseName.objects.get(id=id)
        course.course_name =course_name
        course.save()
        return redirect("app:course")
    return redirect("app:course")

@login_required
def staff_details(request):
    staff = StaffDetails.objects.all()
    df = pd.DataFrame(list(staff.values()))
    df.drop(["user_id","photo"],axis=1,inplace=True)
    file_path = "static/media/data/staff_details.xlsx"
    directory = os.path.dirname(file_path)

    if not os.path.exists(directory):
        os.makedirs(directory)

    df.to_excel(file_path, index=False)
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='staff_details.xlsx')

@login_required
def all_staff(request):
    staff = StaffDetails.objects.all()
    content = {
        "staff":staff
    }
    return render(request,"allStaffs.html",content)



@login_required
def WorkView(request):
    order_map = {'ten_to_twelve': 0, 'one_to_three': 1, 'three_to_five': 2}
    works = []

    if request.method == "POST":
        staff_id_raw = request.POST.get("staff_id")
        selected_staff_id = int(staff_id_raw) if staff_id_raw and staff_id_raw.isdigit() else None
        selected_date_range = request.POST.get("dateRange")

        date_start = None
        date_end = None
        if selected_date_range:
            try:
                if " to " in selected_date_range:
                    date_start_str, date_end_str = selected_date_range.split(" to ")
                    date_start = datetime.strptime(date_start_str.strip(), "%d-%m-%Y").date()
                    date_end = datetime.strptime(date_end_str.strip(), "%d-%m-%Y").date()
                else:
                    date_start = datetime.strptime(selected_date_range.strip(), "%d-%m-%Y").date()
                    date_end = date_start
            except ValueError:
                date_start = datetime.now().date()
                date_end = date_start

            if selected_staff_id is not None:
                works = sorted(
                    Work.objects.filter(user_id=selected_staff_id, date__range=(date_start, date_end)).order_by("date"),
                    key=lambda w: order_map.get(w.time_slot, 99)
                )
            else:
                works = sorted(
                    Work.objects.filter(date__range=(date_start, date_end)).order_by("date"),
                    key=lambda w: order_map.get(w.time_slot, 99)
                )

            context = {
                'works': works,
                'staff_list': User.objects.filter(is_staff=True, is_superuser=False),
                'selected_staff_id': selected_staff_id,
                'selected_date_range': selected_date_range
            }

            if works:
                wb = Workbook()
                ws = wb.active
                ws.title = "Staff Work"
                ws.append(["Date", "Staff Name", "Time Slot", "Description"])
                for work in works:
                    if work.time_slot == "ten_to_twelve":
                        time_slot = "10:00 AM to 12:00 PM"
                    elif work.time_slot == "one_to_three":
                        time_slot = "1:00 PM to 3:00 PM"
                    else:
                        time_slot = "3:00 PM to 5:00 PM"
                    ws.append([
                        work.date.strftime("%d-%m-%Y"),
                        work.user.username,
                        time_slot,
                        work.work
                    ])
                excel_dir = os.path.join(settings.BASE_DIR, 'static', 'media', 'staffworkupdation')
                os.makedirs(excel_dir, exist_ok=True)
                file_path = os.path.join(excel_dir, "staff_work_updation.xlsx")

                if os.path.exists(file_path):
                    os.remove(file_path)

                wb.save(file_path)

            return render(request, "work_updation.html", context)

        else:
            if selected_staff_id is not None:
                works = sorted(
                    Work.objects.filter(user_id=selected_staff_id, date__month=datetime.now().month),
                    key=lambda w: order_map.get(w.time_slot, 99)
                )
            else:
                works = sorted(
                    Work.objects.filter(date__month=datetime.now().month),
                    key=lambda w: order_map.get(w.time_slot, 99)    
                )

            context = {
                'works': works,
                'staff_list': User.objects.filter(is_staff=True, is_superuser=False),
                'selected_staff_id': selected_staff_id,
                'selected_date_range': selected_date_range
            }

            if works:
                wb = Workbook()
                ws = wb.active
                ws.title = "Staff Work"
                ws.append(["Date", "Staff Name", "Time Slot", "Description"])
                for work in works:
                    if work.time_slot == "ten_to_twelve":
                        time_slot = "10:00 AM to 12:00 PM"
                    elif work.time_slot == "one_to_three":
                        time_slot = "1:00 PM to 3:00 PM"
                    else:
                        time_slot = "3:00 PM to 5:00 PM"
                    
                    ws.append([
                        work.date.strftime("%d-%m-%Y"),
                        work.user.username,
                        time_slot,
                        work.work
                    ])
                excel_dir = os.path.join(settings.BASE_DIR, 'static', 'media', 'staffworkupdation')
                os.makedirs(excel_dir, exist_ok=True)
                file_path = os.path.join(excel_dir, "staff_work_updation.xlsx")

                if os.path.exists(file_path):
                    os.remove(file_path)

                wb.save(file_path)

            return render(request, "work_updation.html", context)

    works = sorted(
        Work.objects.filter(date__month=datetime.now().month),
        key=lambda w: order_map.get(w.time_slot, 99)
    )
    

    if works:
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff Work"
        ws.append(["Date", "Staff Name", "Time Slot", "Description"])
        for work in works:
            if work.time_slot == "ten_to_twelve":
                time_slot = "10:00 AM to 12:00 PM"
            elif work.time_slot == "one_to_three":
                time_slot = "1:00 PM to 3:00 PM"
            else:
                time_slot = "3:00 PM to 5:00 PM"
            
            ws.append([
                work.date.strftime("%d-%m-%Y"),
                work.user.username,
                time_slot,
                work.work
            ])
        excel_dir = os.path.join(settings.BASE_DIR, 'static', 'media', 'staffworkupdation')
        os.makedirs(excel_dir, exist_ok=True)
        file_path = os.path.join(excel_dir, "staff_work_updation.xlsx")

        if os.path.exists(file_path):
            os.remove(file_path)

        wb.save(file_path)
    context = {
        'works': works,
        'staff_list': User.objects.filter(is_staff=True, is_superuser=False),
        "selected_date_range": f"{date.today().replace(day=1).strftime('%d-%m-%Y')} to {date.today().strftime('%d-%m-%Y')}"
    }

    return render(request, "work_updation.html", context)



def attendance(request):
    if request.method == "POST":
        search_by = request.POST.get("search_by")
        search_query = request.POST.get("search")
        if search_by == "name":
            students = Student.objects.filter(certificate_id__isnull=True).filter(name__icontains=search_query)
        elif search_by == "phone":
            students = Student.objects.filter(certificate_id__isnull=True).filter(phone__icontains=search_query)
        elif search_by == "email":
            students = Student.objects.filter(certificate_id__isnull=True).filter(email__icontains=search_query)
        elif search_by == "registration_no":
            students = Student.objects.filter(certificate_id__isnull=True).filter(registration_no__icontains=search_query)
        content = {
            "students": students
        }
        return render(request,"attendance.html",content)
    content = {
        "students": Student.objects.all()
    }
    return render(request, "attendance.html",content)


def download_attendance(request):
    if request.method == 'POST':
        student_id = request.POST.get('student_id')

        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return HttpResponse("Student not found", status=404)

        records = Attendance.objects.filter(student=student).order_by('-date')

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Records"

        ws.append(['Date', 'Status', 'Recorded By'])

        for record in records:
            ws.append([
                record.date.strftime('%Y-%m-%d') if record.date else '',
                record.status,
                record.user.username
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={student.name}_attendance.xlsx'

  
        wb.save(response)

        return response



def custom_404(request):
    return render(request, "404.html", status=404)

def custom_500(request):
    return render(request, "500.html", status=500)
