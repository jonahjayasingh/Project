import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import os

def create_attendance_excel(attendance, faculty_name):
    # Prepare attendance data
    attendances = [{
        "Name": a.student.name,
        "Date": a.date.strftime('%Y-%m-%d'),
        "Status": a.status.title()
    } for a in attendance]

    # Create DataFrame
    df = pd.DataFrame(attendances)

    # Pivot safely, handling duplicates using 'first' occurrence
    attendance_pivoted = df.pivot_table(
        index="Name",
        columns="Date",
        values="Status",
        aggfunc="first"  # Use 'first', 'last', 'max', or a custom function as needed
    )

    # Create directory if not exists
    base_dir = "certificategenrator/static/media/attendance"
    os.makedirs(base_dir, exist_ok=True)
    excel_path = os.path.join(base_dir, f"{faculty_name}_attendance.xlsx")

    # Export to Excel
    attendance_pivoted.to_excel(excel_path)

    # Load workbook for formatting
    wb = load_workbook(excel_path)
    ws = wb.active

    # Freeze header and first column
    ws.freeze_panes = "B2"

    # Style header
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Center-align all data cells & adjust row height
    for row in ws.iter_rows(min_row=2):  # Skip header row
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[cell.row].height = 22

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 5

    if os.path.exists(excel_path):
        os.remove(excel_path)
    # Save workbook
    wb.save(excel_path)

    # Return the file name for reference
    return os.path.basename(excel_path)
