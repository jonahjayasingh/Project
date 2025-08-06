import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
import re

def create_work_excel(work_data, name):
    # Convert list of objects to list of dicts
    works = [{
        "date": work.date,
        "time_slot": work.time_slot,
        "work": work.work
    } for work in work_data]

    # Create DataFrame
    df = pd.DataFrame(works)
    if df.empty:
        raise ValueError("No work data provided.")

    # Parse date and extract weekday
    df["date"] = pd.to_datetime(df["date"])
    df["weekday"] = df["date"].dt.strftime("%A")
    df["date"] = df["date"].dt.strftime("%Y-%m-%d")
    df = df[["date", "weekday", "time_slot", "work"]]

    # Group work items by date and time slot
    grouped = df.groupby(['date', 'time_slot'])['work'].apply(list).unstack(fill_value=[])

    # Order time slots
    time_order = ['ten_to_twelve', 'one_to_three', 'three_to_five']
    grouped = grouped[[col for col in time_order if col in grouped.columns]]

    # Normalize row lengths
    max_len = max(grouped.applymap(len).max().max(), 1)
    for col in grouped.columns:
        grouped[col] = grouped[col].apply(lambda x: x + [''] * (max_len - len(x)))

    # Flatten grouped data
    expanded_df = pd.DataFrame({
        col: sum(grouped[col].tolist(), [])
        for col in grouped.columns
    })
    dates = sum([[date] * max_len for date in grouped.index], [])
    expanded_df.index = pd.Index(dates, name='Date')

    # Add weekday column
    final_df = expanded_df.reset_index().copy()
    final_df['Weekday'] = pd.to_datetime(final_df['Date']).dt.strftime('%A')

    # Reorder columns
    cols = ['Date', 'Weekday'] + [col for col in final_df.columns if col not in ['Date', 'Weekday']]
    final_df = final_df[cols]

    # Remove empty rows
    final_df = final_df[~(final_df.iloc[:, 2:] == '').all(axis=1)]

    # Rename time slots for readability
    final_df.rename(columns={
        'ten_to_twelve': '10AM–12PM',
        'one_to_three': '1PM–3PM',
        'three_to_five': '3PM–5PM'
    }, inplace=True)

    # Sort by Date
    final_df['Date'] = pd.to_datetime(final_df['Date'])
    final_df.sort_values('Date', inplace=True)
    final_df['Date'] = final_df['Date'].dt.strftime('%Y-%m-%d')

    # Prepare output path
    work_dir = "static/media/work"
    os.makedirs(work_dir, exist_ok=True)

    # Sanitize filename
    safe_name = re.sub(r'[^\w\-_\. ]', '_', name)
    excel_path = os.path.join(work_dir, f"{safe_name}_workupdate.xlsx")

    # Delete existing file if it exists
    if os.path.exists(excel_path):
        os.remove(excel_path)

    # Save to Excel
    final_df.to_excel(excel_path, index=False, engine='openpyxl')

    # Format Excel
    wb = load_workbook(excel_path)
    ws = wb.active

    # Merge Date and Weekday columns
    start_row = 2
    current_date = ws['A2'].value
    merge_start = start_row

    for row in range(start_row + 1, ws.max_row + 1):
        cell_date = ws[f'A{row}'].value
        if cell_date != current_date:
            if merge_start != row - 1:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=row - 1, end_column=1)
                ws[f'A{merge_start}'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=merge_start, start_column=2, end_row=row - 1, end_column=2)
                ws[f'B{merge_start}'].alignment = Alignment(horizontal='center', vertical='center')
            current_date = cell_date
            merge_start = row

    # Final merge for last group
    if merge_start != ws.max_row:
        ws.merge_cells(start_row=merge_start, start_column=1, end_row=ws.max_row, end_column=1)
        ws[f'A{merge_start}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=merge_start, start_column=2, end_row=ws.max_row, end_column=2)
        ws[f'B{merge_start}'].alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2
    if os.path.exists(excel_path):
        os.remove(excel_path)
    wb.save(excel_path)
    return os.path.basename(excel_path)
